import io
import logging
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, send_file
from flask_cors import cross_origin
from flask_jwt_extended import jwt_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text, func
import requests
from openpyxl.drawing.image import Image

from src.config.database import db
from src.models.catalogos_bench import StRepuestoCompatibilidad, ModeloVersion, ModeloComercial, ClienteCanal, Canal, \
    StCliente, ClienteCanalModelo
from src.models.productos import Producto
from src.models.users import Empresa

bench_rep = Blueprint('routes_bench_rep', __name__)
logger = logging.getLogger(__name__)


@bench_rep.route('/modelos_comerciales_por_marca', methods=['GET'])
def get_modelos_comerciales_por_marca():
    codigo_marca = request.args.get('codigo_marca')

    if not codigo_marca:
        return jsonify({"error": "Parámetro 'codigo_marca' obligatorio"}), 400

    try:
        sql = text("""
            SELECT 
                mc.CODIGO_MODELO_COMERCIAL,
                mc.NOMBRE_MODELO,
                mc.ANIO_MODELO,
                img.PATH_IMAGEN
            FROM STOCK.ST_MODELO_COMERCIAL mc
            LEFT JOIN (
                SELECT 
                    mv.CODIGO_MODELO_COMERCIAL,
                    img.PATH_IMAGEN,
                    ROW_NUMBER() OVER (
                        PARTITION BY mv.CODIGO_MODELO_COMERCIAL 
                        ORDER BY mv.CODIGO_MODELO_VERSION
                    ) AS rn
                FROM STOCK.ST_MODELO_VERSION mv
                JOIN STOCK.ST_IMAGENES img 
                    ON mv.CODIGO_IMAGEN = img.CODIGO_IMAGEN
                WHERE mv.CODIGO_MARCA = :codigo_marca
            ) img 
            ON img.CODIGO_MODELO_COMERCIAL = mc.CODIGO_MODELO_COMERCIAL AND img.rn = 1
            WHERE mc.CODIGO_MARCA = :codigo_marca
        """)
        resultados = db.session.execute(sql, {"codigo_marca": codigo_marca}).mappings().all()

        salida = [
            {
                "codigo_modelo_comercial": row["codigo_modelo_comercial"],
                "nombre_modelo": row["nombre_modelo"],
                "anio_modelo": row["anio_modelo"],
                "path_imagen": row["path_imagen"]
            } for row in resultados
        ]
        return jsonify(salida), 200

    except Exception as e:
        return jsonify({"error": "Error interno", "detalle": str(e)}), 500

@bench_rep.route('/repuestos_compatibles_por_modelo', methods=['GET'])
def repuestos_compatibles_por_modelo():
    codigo_marca = request.args.get('codigo_marca')
    codigo_modelo_comercial = request.args.get('codigo_modelo_comercial')

    if not codigo_modelo_comercial or not codigo_marca:
        return jsonify({"error": "Parámetros 'codigo_marca' y 'codigo_modelo_comercial' son obligatorios"}), 400

    sql = text("""
        SELECT DISTINCT
            mv.CODIGO_MODELO_VERSION,
            mv.NOMBRE_MODELO_VERSION,
            rc.COD_PRODUCTO,
            p.NOMBRE AS nombre_producto,
            rc.EMPRESA,
            e.NOMBRE AS nombre_empresa,
            rc.NIVEL_CONFIANZA,
            rc.VALIDADO_POR,
            rc.ORIGEN_VALIDACION,
            rc.FECHA_VALIDACION,
            rc.COMENTARIOS_TECNICOS,
            mc.NOMBRE_MODELO AS nombre_modelo_comercial,
            mc.ANIO_MODELO,
            p.COD_ITEM_CAT1,
            mi.NOMBRE AS nombre_item,
            img.PATH_IMAGEN AS path_imagen,
            mar.NOMBRE_MARCA AS nombre_marca
        FROM STOCK.ST_MODELO_VERSION mv
            JOIN STOCK.ST_REPUESTO_COMPATIBILIDAD rc
                ON rc.CODIGO_MODELO_VERSION = mv.CODIGO_MODELO_VERSION   
            JOIN STOCK.PRODUCTO p
                ON p.COD_PRODUCTO = rc.COD_PRODUCTO
                AND p.EMPRESA = rc.EMPRESA
            LEFT JOIN STOCK.TG_MODELO_ITEM mi
                ON p.COD_ITEM_CAT1 = mi.COD_ITEM
                AND p.COD_MODELO_CAT1 = mi.COD_MODELO
                AND p.EMPRESA = mi.EMPRESA
            JOIN STOCK.EMPRESA e
                ON rc.EMPRESA = e.EMPRESA
            JOIN STOCK.ST_MODELO_COMERCIAL mc
                ON mc.CODIGO_MODELO_COMERCIAL = mv.CODIGO_MODELO_COMERCIAL
            JOIN STOCK.ST_IMAGENES img
                ON mv.CODIGO_IMAGEN = img.CODIGO_IMAGEN
            JOIN STOCK.ST_MARCA mar
                ON mv.CODIGO_MARCA = mar.CODIGO_MARCA
        WHERE mv.CODIGO_MODELO_COMERCIAL = :codigo_modelo_comercial
          AND mv.CODIGO_MARCA = :codigo_marca
          AND p.COD_MODELO_CAT1 = 'PRO3'
        ORDER BY 
            CASE WHEN rc.NIVEL_CONFIANZA = 100 THEN 0 ELSE 1 END,
            rc.NIVEL_CONFIANZA DESC
    """)

    try:
        resultados = db.session.execute(sql, {
            "codigo_modelo_comercial": codigo_modelo_comercial,
            "codigo_marca": codigo_marca
        }).mappings().all()

        salida = [dict(row) for row in resultados]

        from collections import defaultdict
        agrupados = defaultdict(list)
        for row in salida:
            nombre_item = row.pop("nombre_item") or "Sin Categoría"
            agrupados[nombre_item].append(row)

        return jsonify(agrupados), 200

    except Exception as e:
        print(f"[ERROR] repuestos_compatibles_por_modelo: {str(e)}")
        return jsonify({"error": "Error interno", "detalle": str(e)}), 500


#-----------------------------------------------------------------------------------------------------------------------
#-------------------------------------------ENDPOINT PARA EXPORTAR EL BENCH REPUESTOS ----------------------------------
def descargar_imagen_openpyxl(url, width=200, height=150):
    try:
        if url:
            res = requests.get(url, timeout=2)
            if res.ok:
                img = Image(io.BytesIO(res.content))
                img.width = width
                img.height = height
                return img
    except Exception as e:
        print(f"[IMG] Error al descargar: {url} -> {e}")
    return None

@bench_rep.route('/exportar_modelos_compatibles_xlsx', methods=["POST"])
@jwt_required()
@cross_origin()
def exportar_modelos_compatibles_xlsx():
    try:
        from collections import defaultdict

        data = request.get_json()
        modelos = data.get("modelos", {})
        nombre_modelo = data.get("nombre_modelo", "MODELO DESCONOCIDO")
        nombre_marca = data.get("nombre_marca", "SN")

        if not modelos:
            return jsonify({"error": "No hay modelos compatibles para exportar"}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Modelos Compatibles"

        rojo = "B22222"
        blanco = "FFFFFF"
        gris = "808080"

        gray_border = Border(
            left=Side(border_style="thin", color=gris),
            right=Side(border_style="thin", color=gris),
            top=Side(border_style="thin", color=gris),
            bottom=Side(border_style="thin", color=gris)
        )

        fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)

        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
        cell = ws.cell(row=1, column=3, value="REPUESTOS COMPATIBLES")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.border = gray_border

        # Subtítulo (Marca y Modelo)
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
        cell = ws.cell(row=2, column=3, value=f"{nombre_marca.upper()} - {nombre_modelo.upper()}")

        cell.font = font_white
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

        # Imagen (fila 4)
        imagen_url = next(iter(modelos.values()))[0].get("path_imagen", None)
        if imagen_url:
            try:
                res = requests.get(imagen_url, timeout=3)
                if res.ok:
                    img = Image(io.BytesIO(res.content))
                    img.width = 250
                    img.height = 200
                    ws.add_image(img, "D4")
            except Exception as e:
                print("[ERROR IMG]", e)

        current_row = 15

        columnas = ["ITEM", "COD_PRODUCTO", "NOMBRE_PRODUCTO", "CONFIANZA", "VALIDADO POR", "ORIGEN", "COMENTARIOS"]
        for col_idx, nombre_col in enumerate(columnas, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=nombre_col)
            cell.border = gray_border
            cell.font = Font(bold=True, color=blanco)
            cell.fill = PatternFill(start_color=rojo, end_color=rojo, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
        current_row += 1

        sql_items = text("""
            SELECT DISTINCT p.cod_item_cat1, i.nombre
            FROM producto p
            JOIN stock.tg_modelo_item i ON i.cod_item = p.cod_item_cat1
            WHERE i.cod_modelo = 'PRO3'
        """)
        mapeo_items = {
            row['cod_item_cat1']: row['nombre']
            for row in db.session.execute(sql_items).mappings()
        }

        for categoria, repuestos in modelos.items():
            nombre_item = mapeo_items.get(categoria, "SIN CATEGORÍA").upper()
            inicio_categoria_row = current_row

            for rep in repuestos:
                confianza = rep.get("nivel_confianza", 0)
                if confianza == 100:
                    conf_label, color = "SI", "2e7d32"
                else:
                    conf_label, color = "NO", "d32f2f"

                fila = [
                    None,
                    rep.get("cod_producto", ""),
                    rep.get("nombre_producto", ""),
                    f"{confianza}% – {conf_label}",
                    rep.get("validado_por", ""),
                    rep.get("origen_validacion", ""),
                    rep.get("comentarios_tecnicos", "")
                ]

                for col_idx, valor in enumerate(fila, start=1):
                    c = ws.cell(row=current_row, column=col_idx, value=valor)
                    c.border = gray_border
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    if col_idx == 4:
                        c.font = Font(bold=True, color=color)

                current_row += 1

            # Combinar celdas verticalmente para la categoría en columna 1
            ws.merge_cells(start_row=inicio_categoria_row, end_row=current_row - 1, start_column=1, end_column=1)
            cell_merged = ws.cell(row=inicio_categoria_row, column=1, value=nombre_item)
            cell_merged.alignment = Alignment(horizontal="center", vertical="center")
            cell_merged.font = Font(bold=True)
            cell_merged.border = gray_border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="repuestos_compatibles.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bench_rep.route('/insert_repuesto_compatibilidad', methods=['POST'])
@jwt_required()
def insert_repuesto_compatibilidad():
    try:
        data = request.get_json()

        # Validación de campos obligatorios
        campos_requeridos = [
            'es_compatible', 'validado_por', 'fecha_validacion',
            'nivel_confianza', 'origen_validacion',
            'cod_producto', 'empresa', 'codigo_cliente_canal',
            'codigo_modelo_version', 'codigo_mod_vers_repuesto'
        ]

        faltantes = [campo for campo in campos_requeridos if campo not in data or data[campo] is None]
        if faltantes:
            return jsonify({"error": f"Faltan campos requeridos: {', '.join(faltantes)}"}), 400

        # Conversión de fecha
        try:
            fecha_val = datetime.strptime(data['fecha_validacion'], "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Fecha inválida. Formato requerido: YYYY-MM-DD"}), 400

        # Validación de duplicado antes de insertar (por clave única)
        existente = db.session.query(StRepuestoCompatibilidad).filter_by(
            cod_producto=data['cod_producto'],
            empresa=int(data['empresa']),
            codigo_cliente_canal=int(data['codigo_cliente_canal']),
            codigo_modelo_version=int(data['codigo_modelo_version']),
            codigo_mod_vers_repuesto=int(data['codigo_mod_vers_repuesto'])
        ).first()

        if existente:
            return jsonify({
                "error": "Ya existe una compatibilidad registrada para este modelo, repuesto y cliente-canal"
            }), 409  # Código 409: Conflict

        # Crear nuevo registro
        nuevo = StRepuestoCompatibilidad(
            es_compatible=int(data['es_compatible']),
            validado_por=data['validado_por'],
            fecha_validacion=fecha_val,
            nivel_confianza=int(data['nivel_confianza']),
            origen_validacion=data['origen_validacion'],
            comentarios_tecnicos=data.get('comentarios_tecnicos'),
            cod_producto=data['cod_producto'],
            empresa=int(data['empresa']),
            codigo_cliente_canal=int(data['codigo_cliente_canal']),
            codigo_modelo_version=int(data['codigo_modelo_version']),
            codigo_mod_vers_repuesto=int(data['codigo_mod_vers_repuesto']),
        )

        db.session.add(nuevo)
        db.session.commit()

        return jsonify({
            "message": "Compatibilidad insertada correctamente",
            "id": nuevo.codigo_compatibilidad
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error al insertar", "detalle": str(e)}), 500


#------------------------------------------------------------ INSERT UNIFICADO ST_CLIENTE_CANAL_MODELO Y ST_REPUESTO COMPATIBILIDAD
@bench_rep.route('/insert_cliente_canal_repuesto_compatibilidad', methods=["POST"])
@jwt_required()
def insert_cliente_canal_repuesto_compatibilidad():
    try:
        data = request.get_json()

        campos_requeridos = [
            "codigo_modelo_version", "codigo_mod_vers_repuesto", "codigo_cliente_canal",
            "empresa", "cod_producto", "estado", "fecha_asignacion",
            "es_compatible", "validado_por", "fecha_validacion", "nivel_confianza", "origen_validacion"
        ]
        faltantes = [campo for campo in campos_requeridos if campo not in data or data[campo] in [None, ""]]
        if faltantes:
            return jsonify({"error": f"Faltan campos requeridos: {', '.join(faltantes)}"}), 400

        codigo_modelo_version = int(data["codigo_modelo_version"])
        codigo_cliente_canal = int(data["codigo_cliente_canal"])
        empresa = int(data["empresa"])
        cod_producto = data["cod_producto"]
        codigo_mod_vers_repuesto = int(data["codigo_mod_vers_repuesto"])
        estado = data["estado"]
        estado_c = (
            1 if str(estado).strip().lower() == "activo" else
            0 if str(estado).strip().lower() == "inactivo" else
            int(estado) if str(estado).isdigit() else None
        )

        fecha_asignacion = datetime.strptime(data["fecha_asignacion"], '%Y-%m-%d')
        fecha_validacion = datetime.strptime(data["fecha_validacion"], '%Y-%m-%d')

        if not db.session.query(ModeloVersion).filter_by(codigo_modelo_version=codigo_modelo_version).first():
            return jsonify({"error": "Modelo versión no encontrado"}), 404
        if not db.session.query(ClienteCanal).filter_by(codigo_cliente_canal=codigo_cliente_canal).first():
            return jsonify({"error": "Cliente canal no encontrado"}), 404
        if not db.session.query(Producto).filter_by(cod_producto=cod_producto, empresa=empresa).first():
            return jsonify({"error": "Producto no encontrado"}), 404

        existente_ccm = db.session.query(ClienteCanalModelo).filter_by(
            codigo_modelo_version=codigo_modelo_version,
            codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
            codigo_cliente_canal=codigo_cliente_canal,
            empresa=empresa,
            cod_producto=cod_producto
        ).first()

        if not existente_ccm:
            nuevo_ccm = ClienteCanalModelo(
                codigo_modelo_version=codigo_modelo_version,
                codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
                codigo_cliente_canal=codigo_cliente_canal,
                empresa=empresa,
                cod_producto=cod_producto,
                estado=estado_c,
                fecha_asignacion=fecha_asignacion
            )
            db.session.add(nuevo_ccm)

        existente_rep = db.session.query(StRepuestoCompatibilidad).filter_by(
            codigo_modelo_version=codigo_modelo_version,
            codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
            codigo_cliente_canal=codigo_cliente_canal,
            cod_producto=cod_producto,
            empresa=empresa
        ).first()

        if existente_rep:
            return jsonify({
                "error": "Ya existe una compatibilidad registrada para este modelo, repuesto y cliente-canal"
            }), 409

        nuevo_rc = StRepuestoCompatibilidad(
            es_compatible=int(data['es_compatible']),
            validado_por=data['validado_por'],
            fecha_validacion=fecha_validacion,
            nivel_confianza=int(data['nivel_confianza']),
            origen_validacion=data['origen_validacion'],
            comentarios_tecnicos=data.get('comentarios_tecnicos'),
            cod_producto=cod_producto,
            empresa=empresa,
            codigo_cliente_canal=codigo_cliente_canal,
            codigo_modelo_version=codigo_modelo_version,
            codigo_mod_vers_repuesto=codigo_mod_vers_repuesto
        )
        db.session.add(nuevo_rc)
        db.session.commit()

        return jsonify({
            "message": "Registro insertado correctamente",
            "id": nuevo_rc.codigo_compatibilidad
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error interno", "detalle": str(e)}), 500

@bench_rep.route('/get_cliente_canal_modelo_compatibilidad', methods=['GET'])
@jwt_required()
def get_cliente_canal_modelo_compatibilidad():
    try:
        registros = db.session.query(StRepuestoCompatibilidad).all()
        salida = []

        for r in registros:
            # Producto
            producto = db.session.query(Producto).filter_by(cod_producto=r.cod_producto, empresa=r.empresa).first()

            empresa = db.session.query(Empresa).filter_by(empresa=r.empresa).first()

            modelo_version = db.session.query(ModeloVersion).filter_by(
                codigo_modelo_version=r.codigo_modelo_version).first()
            modelo_comercial = None
            if modelo_version:
                modelo_comercial = db.session.query(ModeloComercial).filter_by(
                    codigo_modelo_comercial=modelo_version.codigo_modelo_comercial
                ).first()

            cliente_canal = db.session.query(ClienteCanal).filter_by(
                codigo_cliente_canal=r.codigo_cliente_canal,
                cod_producto=r.cod_producto,
                empresa=r.empresa,
                codigo_mod_vers_repuesto=r.codigo_mod_vers_repuesto
            ).first()
            canal = None
            cliente = None
            if cliente_canal:
                canal = db.session.query(Canal).filter_by(codigo_canal=cliente_canal.codigo_canal).first()
                cliente = db.session.query(StCliente).filter_by(codigo_cliente=cliente_canal.codigo_cliente).first()

            cliente_canal_modelo = db.session.query(ClienteCanalModelo).filter_by(
                codigo_modelo_version=r.codigo_modelo_version,
                codigo_cliente_canal=r.codigo_cliente_canal,
                empresa=r.empresa,
                cod_producto=r.cod_producto,
                codigo_mod_vers_repuesto=r.codigo_mod_vers_repuesto
            ).first()

            salida.append({
                "codigo_compatibilidad": r.codigo_compatibilidad,
                "es_compatible": r.es_compatible,
                "validado_por": r.validado_por,
                "fecha_validacion": r.fecha_validacion.isoformat(),
                "fecha_asignacion": cliente_canal_modelo.fecha_asignacion.isoformat() if cliente_canal_modelo and cliente_canal_modelo.fecha_asignacion else None,
                "estado": cliente_canal_modelo.estado if cliente_canal_modelo else None,
                "nivel_confianza": r.nivel_confianza,
                "origen_validacion": r.origen_validacion,
                "comentarios_tecnicos": r.comentarios_tecnicos,
                "cod_producto": r.cod_producto,
                "nombre_producto": producto.nombre if producto else None,
                "empresa": r.empresa,
                "nombre_empresa": empresa.nombre if empresa else None,
                "codigo_cliente_canal": r.codigo_cliente_canal,
                "nombre_canal": canal.nombre_canal if canal else None,
                "nombre_cliente": cliente.nombre_cliente if cliente else None,
                "codigo_modelo_version": r.codigo_modelo_version,
                "anio_modelo_comercial": modelo_comercial.anio_modelo if modelo_comercial else None,
                "nombre_modelo_version": modelo_version.nombre_modelo_version if modelo_version else None,
                "nombre_modelo_comercial": modelo_comercial.nombre_modelo if modelo_comercial else None,
                "codigo_mod_vers_repuesto": r.codigo_mod_vers_repuesto
            })
        return jsonify(salida), 200

    except Exception as e:
        return jsonify({"error": "Error al obtener datos", "detalle": str(e)}), 500


@bench_rep.route('/update_cliente_canal_repuesto_compatibilidad/<int:codigo_cliente_canal>', methods=["PUT"])
@jwt_required()
def update_cliente_canal_repuesto_compatibilidad(codigo_cliente_canal):
    try:
        data = request.get_json()

        campos_requeridos = [
            "codigo_modelo_version", "codigo_mod_vers_repuesto", "cod_producto", "empresa",
            "estado", "fecha_asignacion", "es_compatible", "validado_por",
            "fecha_validacion", "nivel_confianza", "origen_validacion"
        ]
        faltantes = [campo for campo in campos_requeridos if campo not in data or data[campo] in [None, ""]]
        if faltantes:
            return jsonify({"error": f"Faltan campos requeridos: {', '.join(faltantes)}"}), 400

        codigo_modelo_version = int(data["codigo_modelo_version"])
        codigo_mod_vers_repuesto = int(data["codigo_mod_vers_repuesto"])
        cod_producto = data["cod_producto"]
        empresa = int(data["empresa"])
        estado = int(data["estado"])
        es_compatible = int(data["es_compatible"])
        nivel_confianza = int(data["nivel_confianza"])
        validado_por = data["validado_por"]
        origen_validacion = data["origen_validacion"]
        comentarios_tecnicos = data.get("comentarios_tecnicos")

        fecha_asignacion = datetime.fromisoformat(data["fecha_asignacion"])
        fecha_validacion = datetime.fromisoformat(data["fecha_validacion"])

        ccm = db.session.query(ClienteCanalModelo).filter_by(
            codigo_cliente_canal=codigo_cliente_canal,
            codigo_modelo_version=codigo_modelo_version,
            codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
            cod_producto=cod_producto,
            empresa=empresa
        ).first()

        if not ccm:
            return jsonify({"error": "No se encontró ClienteCanalModelo a actualizar"}), 404

        ccm.estado = estado
        ccm.fecha_asignacion = fecha_asignacion

        rc = db.session.query(StRepuestoCompatibilidad).filter_by(
            codigo_cliente_canal=codigo_cliente_canal,
            codigo_modelo_version=codigo_modelo_version,
            codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
            cod_producto=cod_producto,
            empresa=empresa
        ).first()

        if not rc:
            return jsonify({"error": "No se encontró compatibilidad registrada para actualizar"}), 404

        rc.es_compatible = es_compatible
        rc.validado_por = validado_por
        rc.fecha_validacion = fecha_validacion
        rc.nivel_confianza = nivel_confianza
        rc.origen_validacion = origen_validacion
        rc.comentarios_tecnicos = comentarios_tecnicos

        db.session.commit()

        return jsonify({"message": "Registro actualizado correctamente"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error interno", "detalle": str(e)}), 500


#---------------------------------------------------------------------------
#-------------------------------------ENDPOINT QUE MUESTRA LOS REPUESTOS POR CATÉGORIAS
@bench_rep.route('/repuestos_compatibles', methods=['GET'])
def get_repuestos_compatibles():
    modelo_version = request.args.get('modelo_version', type=int)
    cliente_canal = request.args.get('cliente_canal', type=int)
    codigo_mod_vers_repuesto = request.args.get('codigo_mod_vers_repuesto', type=int)
    empresa = request.args.get('empresa', type=int)

    if not all([modelo_version, cliente_canal, codigo_mod_vers_repuesto, empresa]):
        return jsonify({"error": "Faltan parámetros requeridos"}), 400

    query = text("""
        SELECT
            tmi.NOMBRE AS categoria,
            rc.COD_PRODUCTO,
            p.NOMBRE AS nombre_repuesto,
            rc.NIVEL_CONFIANZA,
            rc.ORIGEN_VALIDACION,
            rc.COMENTARIOS_TECNICOS
        FROM ST_REPUESTO_COMPATIBILIDAD rc
        JOIN PRODUCTO p
            ON p.COD_PRODUCTO = rc.COD_PRODUCTO
           AND p.EMPRESA = rc.EMPRESA
        JOIN COMPUTO.TG_MODELO_ITEM tmi
            ON tmi.COD_ITEM = p.COD_ITEM
           AND tmi.COD_MODELO = p.COD_MODELO
           AND tmi.EMPRESA = p.EMPRESA
        WHERE rc.CODIGO_MODELO_VERSION = :modelo_version
          AND rc.CODIGO_CLIENTE_CANAL = :cliente_canal
          AND rc.CODIGO_MOD_VERS_REPUESTO = :codigo_mod_vers_repuesto
          AND rc.EMPRESA = :empresa          
        ORDER BY categoria, nombre_repuesto
    """)

    result = db.session.execute(query, {
        "modelo_version": modelo_version,
        "cliente_canal": cliente_canal,
        "codigo_mod_vers_repuesto": codigo_mod_vers_repuesto,
        "empresa": empresa
    })

    # Agrupar por categoría
    data = {}
    for row in result:
        categoria = row.categoria
        repuesto = {
            "cod_producto": row.cod_producto,
            "nombre_repuesto": row.nombre_repuesto,
            "nivel_confianza": row.nivel_confianza,
            "origen_validacion": row.origen_validacion,
            "comentarios_tecnicos": row.comentarios_tecnicos
        }
        if categoria not in data:
            data[categoria] = []
        data[categoria].append(repuesto)

    return jsonify(data)


#-------------------------------------- REGISTROS MASIVOS

def excel_date_to_str(value):
    if isinstance(value, int) or isinstance(value, float):
        date = datetime(1899, 12, 30) + timedelta(days=int(value))
        return date.strftime("%Y/%m/%d")
    return value

@bench_rep.route('/insert_repuesto_compatibilidad_masivo', methods=['POST'])
@jwt_required()
def insert_repuesto_compatibilidad_masivo():
    try:
        data = request.get_json().get('repuestos')
        if not isinstance(data, list) or not data:
            return jsonify({"error": "El campo 'repuestos' debe contener una lista de registros"}), 400

        empresa_id = 20
        registros_repuesto = []
        registros_cliente_canal_modelo = []
        errores = []
        duplicados = []

        for i, row in enumerate(data):
            fila = f"Fila {i+1}"
            try:
                campos_obligatorios = [
                    'nombre_modelo_comercial', 'nombre_cliente', 'nombre_canal', 'nombre_producto',
                    'estado', 'fecha_asignacion',
                    'es_compatible', 'validado_por', 'fecha_validacion',
                    'nivel_confianza', 'origen_validacion'
                ]
                for campo in campos_obligatorios:
                    if campo not in row or row[campo] in [None, ""]:
                        raise ValueError(f"{fila}: Falta campo obligatorio: {campo}")

                modelo_version = db.session.query(ModeloVersion).join(ModeloComercial).filter(
                    ModeloComercial.nombre_modelo == row['nombre_modelo_comercial'].strip()
                ).first()

                cliente = db.session.query(StCliente).filter_by(nombre_cliente=row['nombre_cliente']).first()
                canal = db.session.query(Canal).filter_by(nombre_canal=row['nombre_canal']).first()

                producto = db.session.query(Producto).filter(
                    func.upper(func.trim(Producto.nombre)) == row['nombre_producto'].strip().upper()
                ).first()

                if not producto:
                    raise ValueError(f"{fila}: Producto '{row['nombre_producto']}' no encontrado.")

                cliente_canal = db.session.query(ClienteCanal).filter_by(
                    codigo_cliente=cliente.codigo_cliente,
                    codigo_canal=canal.codigo_canal,
                    cod_producto=producto.cod_producto,
                    empresa=empresa_id
                ).first()

                if not cliente_canal:
                    raise ValueError(f"{fila}: No se encontró ClienteCanal con cliente, canal, producto y empresa")

                if not all([modelo_version, cliente, canal, producto]):
                    raise ValueError(f"{fila}: No se encontraron todas las entidades requeridas")

                estado_map = {"ACTIVO": 1, "INACTIVO": 0}
                estado = estado_map.get(str(row['estado']).strip().upper(), row['estado'])
                if str(estado) not in ["1", "0"]:
                    raise ValueError(f"{fila}: Estado inválido (debe ser 1, 0, ACTIVO o INACTIVO)")

                compatible_map = {"SI": 1, "NO": 0}
                es_compatible = compatible_map.get(str(row['es_compatible']).strip().upper(), row['es_compatible'])
                if str(es_compatible) not in ["1", "0"]:
                    raise ValueError(f"{fila}: es_compatible inválido (debe ser 1, 0, SI o NO)")

                nivel_confianza = int(row['nivel_confianza'])
                if not (0 <= nivel_confianza <= 100):
                    raise ValueError(f"{fila}: nivel_confianza fuera de rango (0–100)")

                origen = row['origen_validacion'].strip().upper()
                if origen not in ["INTERNO", "PROVEEDOR", "FABRICANTE"]:
                    raise ValueError(f"{fila}: origen_validacion inválido (INTERNO, PROVEEDOR o FABRICANTE)")

                try:
                    fecha_validacion = datetime.strptime(
                        excel_date_to_str(row['fecha_validacion']), "%Y/%m/%d"
                    )
                    fecha_asignacion = datetime.strptime(
                        excel_date_to_str(row['fecha_asignacion']), "%Y/%m/%d"
                    )
                except ValueError:
                    raise ValueError(f"{fila}: Formato de fecha inválido (usar yyyy/mm/dd)")

                existe_compat = db.session.query(StRepuestoCompatibilidad).filter_by(
                    cod_producto=producto.cod_producto,
                    codigo_cliente_canal=cliente_canal.codigo_cliente_canal,
                    codigo_modelo_version=modelo_version.codigo_modelo_version,
                ).first()

                if existe_compat:
                    duplicados.append(fila)
                    continue
                duplicados.append({
                    "fila": i + 1,
                    "error": f"{fila}: Registro DUPLICADO"
                })

                existe_ccm = db.session.query(ClienteCanalModelo).filter_by(
                    cod_producto=producto.cod_producto,
                    codigo_cliente_canal=cliente_canal.codigo_cliente_canal,
                    codigo_modelo_version=modelo_version.codigo_modelo_version
                ).first()

                if not existe_ccm:
                    nuevo_ccm = ClienteCanalModelo(
                        cod_producto=producto.cod_producto,
                        empresa=empresa_id,
                        codigo_cliente_canal=cliente_canal.codigo_cliente_canal,
                        codigo_modelo_version=modelo_version.codigo_modelo_version,
                        codigo_mod_vers_repuesto=cliente_canal.codigo_mod_vers_repuesto,
                        fecha_asignacion=fecha_asignacion,
                        estado=int(estado)
                    )
                    registros_cliente_canal_modelo.append(nuevo_ccm)

                nuevo_compat = StRepuestoCompatibilidad(
                    es_compatible=int(es_compatible),
                    validado_por=row['validado_por'],
                    fecha_validacion=fecha_validacion,
                    nivel_confianza=nivel_confianza,
                    origen_validacion=origen,
                    comentarios_tecnicos=row.get('comentarios_tecnicos'),
                    cod_producto=producto.cod_producto,
                    codigo_modelo_version=modelo_version.codigo_modelo_version,
                    empresa=empresa_id,
                    codigo_cliente_canal=cliente_canal.codigo_cliente_canal,
                    codigo_mod_vers_repuesto=cliente_canal.codigo_mod_vers_repuesto
                )
                registros_repuesto.append(nuevo_compat)

            except Exception as e:
                errores.append({"fila": fila, "error": str(e)})

        if registros_cliente_canal_modelo:
            db.session.bulk_save_objects(registros_cliente_canal_modelo)
        if registros_repuesto:
            db.session.bulk_save_objects(registros_repuesto)

        db.session.commit()

        return jsonify({
            "insertados": len(registros_repuesto),
            "duplicados": duplicados,
            "errores": errores,
            "message": f"Insertados: {len(registros_repuesto)}, Duplicados: {len(duplicados)}"
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "error": "Error general",
            "detalle": str(e)
        }), 500

@bench_rep.route('/update_cliente_canal_repuesto_compatibilidad_masivo', methods=["PUT"])
@jwt_required()
def update_cliente_canal_repuesto_compatibilidad_masivo():
    try:
        registros = request.get_json()
        if not isinstance(registros, list) or not registros:
            return jsonify({"error": "Se esperaba un array JSON con los registros a actualizar"}), 400

        for i, row in enumerate(registros):
            fila = f"Fila {i + 1}"

            campos_requeridos = [
                "codigo_cliente_canal", "nombre_canal", "nombre_cliente",
                "nombre_producto", "nombre_modelo_comercial",
                "estado", "fecha_asignacion", "es_compatible", "validado_por",
                "fecha_validacion", "nivel_confianza", "origen_validacion"
            ]

            for campo in campos_requeridos:
                if campo not in row or row[campo] in [None, ""]:
                    raise ValueError(f"{fila}: Falta campo obligatorio: {campo}")
            try:
                codigo_cliente_canal = int(row["codigo_cliente_canal"])
                estado = int(row["estado"])
                es_compatible = int(row["es_compatible"])
                validado_por = row["validado_por"].strip()
                nivel_confianza = int(row["nivel_confianza"])
                origen = row["origen_validacion"].strip().upper()
                comentarios_tecnicos = row.get("comentarios_tecnicos")

                if origen not in ["INTERNO", "PROVEEDOR", "FABRICANTE"]:
                    raise ValueError(f"{fila}: origen_validacion inválido")

                def parse_fecha(valor):
                    if isinstance(valor, (int, float)):
                        return datetime(1899, 12, 30) + timedelta(days=int(valor))
                    return datetime.strptime(str(valor).strip(), "%Y/%m/%d")

                fecha_asignacion = parse_fecha(row["fecha_asignacion"])
                fecha_validacion = parse_fecha(row["fecha_validacion"])

                # Resolver canal
                canal = db.session.query(Canal).filter(
                    func.upper(Canal.nombre_canal) == row["nombre_canal"].strip().upper()
                ).first()
                if not canal:
                    raise ValueError(f"{fila}: Canal '{row['nombre_canal']}' no encontrado")

                cliente = db.session.query(StCliente).filter(
                    func.upper(StCliente.nombre_cliente) == row["nombre_cliente"].strip().upper()
                ).first()
                if not cliente:
                    raise ValueError(f"{fila}: Cliente '{row['nombre_cliente']}' no encontrado")

                producto = db.session.query(Producto).filter(
                    func.upper(func.trim(Producto.nombre)) == row["nombre_producto"].strip().upper()
                ).first()
                if not producto:
                    raise ValueError(f"{fila}: Producto '{row['nombre_producto']}' no encontrado")

                cliente_canal = db.session.query(ClienteCanal).filter_by(
                    codigo_cliente=cliente.codigo_cliente,
                    codigo_canal=canal.codigo_canal,
                    cod_producto=producto.cod_producto
                ).first()
                if not cliente_canal:
                    raise ValueError(f"{fila}: No se encontró ClienteCanal con cliente, canal, producto")

                modelo_version = db.session.query(ModeloVersion).join(ModeloComercial).filter(
                    func.upper(ModeloComercial.nombre_modelo) == row["nombre_modelo_comercial"].strip().upper()
                ).first()
                if not modelo_version:
                    raise ValueError(f"{fila}: Modelo comercial '{row['nombre_modelo_comercial']}' no encontrado")

                empresa = cliente_canal.empresa
                cod_producto = cliente_canal.cod_producto
                codigo_mod_vers_repuesto = cliente_canal.codigo_mod_vers_repuesto

                # Buscar y actualizar ST_CLIENTE_CANAL_MODELO
                ccm = db.session.query(ClienteCanalModelo).filter_by(
                    codigo_cliente_canal=codigo_cliente_canal,
                    codigo_modelo_version=modelo_version.codigo_modelo_version,
                    codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
                    cod_producto=cod_producto,
                    empresa=empresa
                ).first()
                if not ccm:
                    raise ValueError(f"{fila}: No se encontró ClienteCanalModelo para actualizar")

                ccm.estado = estado
                ccm.fecha_asignacion = fecha_asignacion

                # Buscar y actualizar ST_REPUESTO_COMPATIBILIDAD
                rc = db.session.query(StRepuestoCompatibilidad).filter_by(
                    codigo_cliente_canal=codigo_cliente_canal,
                    codigo_modelo_version=modelo_version.codigo_modelo_version,
                    codigo_mod_vers_repuesto=codigo_mod_vers_repuesto,
                    cod_producto=cod_producto,
                    empresa=empresa
                ).first()
                if not rc:
                    raise ValueError(f"{fila}: No se encontró RepuestoCompatibilidad para actualizar")

                rc.es_compatible = es_compatible
                rc.validado_por = validado_por
                rc.fecha_validacion = fecha_validacion
                rc.nivel_confianza = nivel_confianza
                rc.origen_validacion = origen
                rc.comentarios_tecnicos = comentarios_tecnicos

            except ValueError as ve:
                raise ve
            except Exception as e:
                raise ValueError(f"{fila}: Error inesperado → {str(e)}")

        db.session.commit()
        return jsonify({"message": f"Se actualizaron correctamente {len(registros)} registros"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "error": "Error al procesar registros",
            "detalle": str(e)
        }), 500