import logging
import re
from flask import Blueprint, jsonify, request
from flask_cors import cross_origin
from flask_jwt_extended import jwt_required
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from sqlalchemy import func, text, Float, or_, and_
from openpyxl import Workbook
from flask import send_file
from collections import defaultdict
from src.config.database import db
from src.models.catalogos_bench import ModeloVersion, Motor, TipoMotor, Chasis, ElectronicaOtros, DimensionPeso, \
    Transmision, ClienteCanal, Segmento, ModeloComercial, Version, Marca, Imagenes
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
import io
import requests

bench_model = Blueprint('routes_bench_model', __name__)
logger = logging.getLogger(__name__)

@bench_model.route('/comparar_modelos', methods=["POST"])
@jwt_required()
def comparar_modelos():
    try:
        data = request.get_json()
        base_id = data.get("modelo_base")
        comparables_ids = data.get("comparables", [])

        if not base_id or not comparables_ids:
            return jsonify({"error": "Se requiere modelo base y al menos un modelo comparable"}), 400

        # FUNCIONES DE EXTRACCIÓN NUMÉRICA
        def extract_kmh(val):
            match = re.search(r'([\d.]+)\s*km', str(val).lower())
            return float(match.group(1)) if match else None

        def extract_litros(val):
            match = re.search(r'([\d.]+)\s*litros?', str(val).lower())
            return float(match.group(1)) if match else None

        def extract_cc(val):
            match = re.search(r'([\d.]+)\s*cc', str(val).lower())
            return float(match.group(1)) if match else None

        def extract_garantia_meses(texto):
            if not texto:
                return None

            texto = str(texto).upper().replace(".", ",")
            partes = re.split(r"/| O | Y |-", texto)

            max_meses = 0

            for parte in partes:
                match_anos = re.search(r'(\d+(?:[.,]\d+)?)\s*AÑOS?', parte)
                if match_anos:
                    meses = float(match_anos.group(1).replace(",", ".")) * 12
                    max_meses = max(max_meses, meses)
                match_meses = re.search(r'(\d+(?:[.,]\d+)?)\s*MESES?', parte)
                if match_meses:
                    meses = float(match_meses.group(1).replace(",", "."))
                    max_meses = max(max_meses, meses)

            return int(max_meses) if max_meses > 0 else None

        def evaluar_pneumatic(cadena):
            if not cadena:
                return None
            cadena = cadena.replace(" ", "")
            match = re.match(r"(\d+)[/-](\d+)[/-]?(\d+)?", cadena)
            if not match:
                return None
            ancho = int(match.group(1))
            relacion = int(match.group(2))
            rin = int(match.group(3)) if match.group(3) else 0
            return 2 * (ancho * (relacion / 100)) + (rin * 25.4)

        def extract_hp(val):
            if not val:
                return None
            val = str(val).lower().replace(",", ".")
            match = re.search(r'(\d+(?:\.\d+)?)\s*hp', val)
            return float(match.group(1)) if match else None

        def extract_nm(val):
            if not val:
                return None
            val = str(val).lower().replace(",", ".")
            match = re.search(r'(\d+(?:\.\d+)?)\s*(nm|bnm)', val)
            return float(match.group(1)) if match else None

        def extract_cambios(val):
            if not val:
                return None
            val = str(val).lower()
            if 'automática' in val:
                return 100
            if 'semi' in val:
                return 50
            match = re.search(r'(\d+)', val)
            return int(match.group(1)) if match else None

        # SEGMENTO
        segmento = db.session.query(Segmento.nombre_segmento).join(ModeloComercial,
            (Segmento.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
            (Segmento.codigo_marca == ModeloComercial.codigo_marca))
        segmento = segmento.filter(ClienteCanal.codigo_mod_vers_repuesto == base_id).first()
        segmento_nombre = segmento[0].lower() if segmento else ""

        mejor_si_menor = set()
        mejor_si_mayor = set()
        mejor_si_diferente = set()

        siempre_mejor_si_mayor = {
            "velocidad_maxima","cilindrada_comercial","caballos_fuerza","capacidad_combustible","peso_seco",
            "torque_maximo","caja_cambios","neumatico_delantero","neumatico_trasero","garantia",
            "ancho_total","longitud_total","altura_total"
        }

        if segmento_nombre == "cross":
            mejor_si_diferente.update({
                "suspension_delantera", "suspension_trasera", "aros_rueda_delantera",
                "aros_rueda_posterior", "frenos_traseros", "frenos_delanteros",
                "tablero", "luces_delanteras", "luces_posteriores",
                "sistema_combustible", "sistema_refrigeracion", "arranque"
            })
        elif segmento_nombre == "scooter":
            mejor_si_menor.update({
                "cilindrada_comercial","peso_seco","caballos_fuerza","torque_maximo",
                "altura_total","ancho_total","longitud_total"
            })
        elif segmento_nombre == "caballito":
            mejor_si_menor.update({
                "cilindrada_comercial", "peso_seco", "caballos_fuerza", "torque_maximo",
                "altura_total", "ancho_total", "longitud_total"
            })
        elif segmento_nombre == "adventure":
            mejor_si_mayor.update({
                "altura_total", "ancho_total", "longitud_total"
            })
        elif segmento_nombre == "utilitaria":
            mejor_si_menor.update({
                "cilindrada_comercial", "altura_total", "ancho_total", "longitud_total"
            })
        elif segmento_nombre == "street":
            mejor_si_menor.update({
                "cilindrada_comercial", "altura_total", "ancho_total", "longitud_total"
            })
        elif segmento_nombre == "ninja":
            mejor_si_menor.update({
                "cilindrada_comercial", "altura_total", "ancho_total", "longitud_total"
            })
        elif segmento_nombre == "deportiva":
            mejor_si_menor.update({
                "altura_total", "ancho_total", "longitud_total",
            })

        mejor_si_mayor.update(siempre_mejor_si_mayor)

        def evaluar_estado(campo, base_val, comp_val):
            campos_diferentes = {
                "suspension_delantera", "suspension_trasera",
                "aros_rueda_delantera", "aros_rueda_posterior",
                "frenos_traseros", "frenos_delanteros",
                "tablero", "luces_delanteras", "luces_posteriores",
                "sistema_combustible", "sistema_refrigeracion", "arranque"
            }

            if not base_val or not comp_val:
                return "diferente"

            if campo in campos_diferentes:
                base = str(base_val).strip().lower()
                comp = str(comp_val).strip().lower()
                return "igual" if base == comp else "diferente"

            extractores = {
                "caballos_fuerza": extract_hp,
                "torque_maximo": extract_nm,
                "velocidad_maxima": extract_kmh,
                "capacidad_combustible": extract_litros,
                "garantia": extract_garantia_meses,
                "cilindrada_comercial": extract_cc,
                "neumatico_delantero": evaluar_pneumatic,
                "neumatico_trasero": evaluar_pneumatic,
                "caja_cambios": extract_cambios
            }

            if campo in extractores:
                base = extractores[campo](base_val)
                comp = extractores[campo](comp_val)
            else:
                try:
                    base = float(base_val)
                    comp = float(comp_val)
                except (TypeError, ValueError):
                    return "diferente"

            if base is None or comp is None:
                return "diferente"

            if isinstance(base, (int, float)) and isinstance(comp, (int, float)):
                if abs(base - comp) < 0.01:
                    return "igual"

            if campo in mejor_si_mayor:
                return "mejor" if comp > base else "peor" if comp < base else "igual"
            if campo in mejor_si_menor:
                return "mejor" if comp < base else "peor" if comp > base else "igual"

            return "igual"

        def cargar_detalles(modelos):
            return {
                "modelo_version": modelos.codigo_modelo_version,
                "nombre_modelo_version": modelos.nombre_modelo_version,
                "motor": db.session.query(Motor).get((modelos.codigo_motor, modelos.codigo_tipo_motor)),
                "tipo_motor": db.session.query(TipoMotor).get(modelos.codigo_tipo_motor),
                "chasis": db.session.query(Chasis).get(modelos.codigo_chasis),
                "electronica": db.session.query(ElectronicaOtros).get(modelos.codigo_electronica),
                "dimensiones": db.session.query(DimensionPeso).get(modelos.codigo_dim_peso),
                "transmision": db.session.query(Transmision).get(modelos.codigo_transmision),
            }

        base_modelo = db.session.query(ModeloVersion).filter_by(codigo_modelo_version=base_id).first()
        if not base_modelo:
            return jsonify({"error": "Modelo base no encontrado"}), 404

        detalles_base = cargar_detalles(base_modelo)
        resultado = []

        for comp_id in comparables_ids:
            modelo = db.session.query(ModeloVersion).filter_by(codigo_modelo_version=comp_id).first()
            if not modelo:
                continue
            detalles_comp = cargar_detalles(modelo)
            mejor_en = {}
            diferente_en = {}

            def comparar(campo, val_base, val_comp, grupo):
                estado = evaluar_estado(campo, val_base, val_comp)
                mejor_en.setdefault(grupo, []).append({
                    "campo": campo,
                    "base": val_base,
                    "comparable": val_comp,
                    "estado": estado
                })
                diferente_en.setdefault(grupo, []).append({
                    "campo": campo,
                    "base": val_base,
                    "comparable": val_comp,
                    "estado": estado
                })

            comparar("cilindrada_comercial",
                     detalles_base["motor"].cilindrada,
                     detalles_comp["motor"].cilindrada, "motor")
            comparar("caballos_fuerza",
                     detalles_base["motor"].caballos_fuerza,
                     detalles_comp["motor"].caballos_fuerza, "motor")
            comparar("torque_maximo",
                     detalles_base["motor"].torque_maximo,
                     detalles_comp["motor"].torque_maximo, "motor")
            comparar("sistema_combustible",
                     detalles_base["motor"].sistema_combustible,
                     detalles_comp["motor"].sistema_combustible, "motor")
            comparar("arranque",
                     detalles_base["motor"].arranque,
                     detalles_comp["motor"].arranque, "motor")
            comparar("sistema_refrigeracion",
                     detalles_base["motor"].sistema_refrigeracion,
                     detalles_comp["motor"].sistema_refrigeracion, "motor")

            comparar("suspension_delantera",
                     detalles_base["chasis"].suspension_delantera,
                     detalles_comp["chasis"].suspension_delantera, "chasis")
            comparar("suspension_trasera",
                     detalles_base["chasis"].suspension_trasera,
                     detalles_comp["chasis"].suspension_trasera, "chasis")
            comparar("aros_rueda_delantera",
                     detalles_base["chasis"].aros_rueda_delantera,
                     detalles_comp["chasis"].aros_rueda_delantera, "chasis")
            comparar("aros_rueda_posterior",
                     detalles_base["chasis"].aros_rueda_posterior,
                     detalles_comp["chasis"].aros_rueda_posterior, "chasis")
            comparar("neumatico_delantero",
                     detalles_base["chasis"].neumatico_delantero,
                     detalles_comp["chasis"].neumatico_delantero, "chasis")
            comparar("neumatico_trasero",
                     detalles_base["chasis"].neumatico_trasero,
                     detalles_comp["chasis"].neumatico_trasero, "chasis")
            comparar("frenos_traseros",
                     detalles_base["chasis"].frenos_traseros,
                     detalles_comp["chasis"].frenos_traseros, "chasis")
            comparar("frenos_delanteros",
                     detalles_base["chasis"].frenos_delanteros,
                     detalles_comp["chasis"].frenos_delanteros, "chasis")

            comparar("altura_total",
                     detalles_base["dimensiones"].altura_total,
                     detalles_comp["dimensiones"].altura_total, "dimensiones")
            comparar("longitud_total",
                     detalles_base["dimensiones"].longitud_total,
                     detalles_comp["dimensiones"].longitud_total, "dimensiones")
            comparar("ancho_total",
                     detalles_base["dimensiones"].ancho_total,
                     detalles_comp["dimensiones"].ancho_total, "dimensiones")
            comparar("peso_seco",
                     detalles_base["dimensiones"].peso_seco,
                     detalles_comp["dimensiones"].peso_seco, "dimensiones")

            comparar("capacidad_combustible",
                     detalles_base["electronica"].capacidad_combustible,
                     detalles_comp["electronica"].capacidad_combustible, "electronica")
            comparar("tablero",
                     detalles_base["electronica"].tablero, detalles_comp["electronica"].tablero, "electronica")
            comparar("luces_delanteras",
                     detalles_base["electronica"].luces_delanteras,
                     detalles_comp["electronica"].luces_delanteras, "electronica")
            comparar("luces_posteriores",
                     detalles_base["electronica"].luces_posteriores,
                     detalles_comp["electronica"].luces_posteriores, "electronica")
            comparar("garantia",
                     detalles_base["electronica"].garantia,
                     detalles_comp["electronica"].garantia, "electronica")
            comparar("velocidad_maxima",
                     detalles_base["electronica"].velocidad_maxima,
                     detalles_comp["electronica"].velocidad_maxima, "electronica")

            comparar("caja_cambios",
                     detalles_base["transmision"].caja_cambios,
                     detalles_comp["transmision"].caja_cambios, "transmision")

            resultado.append({
                "modelo_version": comp_id,
                "nombre_modelo": modelo.nombre_modelo_version,
                "mejor_en": mejor_en,
                "diferente_en": diferente_en
            })
        return jsonify({"base": base_id, "comparables": resultado})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bench_model.route('/get_modelos_por_linea/<int:codigo_linea>', methods=['GET'])
@jwt_required()
@cross_origin()
def get_modelos_por_linea(codigo_linea):
    try:
        resultados = db.session.query(
            ModeloVersion.codigo_modelo_version,
            ModeloVersion.nombre_modelo_version,
            ModeloVersion.anio_modelo_version,
            ModeloVersion.precio_producto_modelo,
            ModeloComercial.nombre_modelo.label('nombre_modelo_comercial'),
            Motor.nombre_motor
        ).join(Segmento, (Segmento.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
                         (Segmento.codigo_marca == ModeloComercial.codigo_marca)) \
         .join(Motor, ModeloVersion.codigo_motor == Motor.codigo_motor) \
         .filter(Segmento.codigo_linea == codigo_linea) \
         .all()

        modelos = [
            {
                "codigo_modelo_version": r[0],
                "nombre_modelo_version": r[1],
                "anio_modelo_version": r[2],
                "precio_producto_modelo": r[3],
                "nombre_modelo_comercial": r[4],
                "nombre_motor": r[5]
            }
            for r in resultados
        ]

        return jsonify(modelos), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bench_model.route('/get_modelos_por_linea_segmento', methods=['GET'])
@jwt_required()
@cross_origin()
def get_modelos_por_linea_segmento():
    try:
        codigo_linea = request.args.get('codigo_linea', type=int)
        nombre_segmento = request.args.get('nombre_segmento', type=str)

        if not codigo_linea or not nombre_segmento:
            return jsonify({"error": "Parámetros 'codigo_linea' y 'nombre_segmento' requeridos"}), 400

        resultados = db.session.query(
            ModeloVersion.codigo_modelo_version,
            ModeloVersion.nombre_modelo_version,
            ModeloVersion.anio_modelo_version,
            ModeloVersion.precio_producto_modelo,
            ModeloComercial.nombre_modelo.label('nombre_modelo_comercial'),
            Motor.nombre_motor,
            TipoMotor.nombre_tipo,
            Version.nombre_version,
            Marca.nombre_marca,
            Imagenes.path_imagen,
        ) \
            .join(ModeloComercial, (ModeloVersion.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
                  (ModeloVersion.codigo_marca == ModeloComercial.codigo_marca)) \
            .join(Segmento, (Segmento.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
                  (Segmento.codigo_marca == ModeloComercial.codigo_marca)) \
            .join(Motor, ModeloVersion.codigo_motor == Motor.codigo_motor) \
            .join(TipoMotor, Motor.codigo_tipo_motor == TipoMotor.codigo_tipo_motor) \
            .join(Version, ModeloVersion.codigo_version == Version.codigo_version) \
            .join(Imagenes, ModeloVersion.codigo_imagen == Imagenes.codigo_imagen) \
            .join(Marca, ModeloVersion.codigo_marca == Marca.codigo_marca) \
            .filter(Segmento.codigo_linea == codigo_linea) \
            .filter(func.upper(Segmento.nombre_segmento) == func.upper(nombre_segmento.strip())) \
            .all()

        modelos = [{
            "codigo_modelo_version": r[0],
            "nombre_modelo_version": r[1],
            "anio_modelo_version": r[2],
            "precio_producto_modelo": r[3],
            "nombre_modelo_comercial": r[4],
            "nombre_motor": r[5],
            "nombre_tipo": r[6],
            "nombre_version": r[7],
            "nombre_marca": r[8],
            "path_imagen": r[9],
        } for r in resultados]

        return jsonify(modelos), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bench_model.route('/get_segmentos_por_linea/<int:codigo_linea>', methods=['GET'])
@jwt_required()
@cross_origin()
def get_segmentos_por_linea(codigo_linea):
    try:
        segmentos = db.session.query(
            Segmento.nombre_segmento,
            db.func.min(Segmento.codigo_segmento).label('codigo_segmento')
        ).join(ModeloComercial, ModeloComercial.codigo_modelo_comercial == Segmento.codigo_modelo_comercial)
        segmentos = segmentos.filter(
            Segmento.codigo_linea == codigo_linea,
            Segmento.estado_segmento == 1,
            ModeloComercial.estado_modelo == 1
        ).group_by(
            Segmento.nombre_segmento
        ).order_by(
            Segmento.nombre_segmento
        ).all()

        return jsonify([
            {
                "codigo_segmento": s.codigo_segmento,
                "nombre_segmento": s.nombre_segmento
            } for s in segmentos
        ]), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bench_model.route('/exportar_comparacion_xlsx', methods=["POST"])
@jwt_required()
@cross_origin()
def exportar_comparacion_xlsx():
    try:
        data = request.get_json()
        resultado = data.get("resultado")
        modelos = data.get("modelos")

        gray_border = Border(
            left=Side(border_style="thin", color="808080"),
            right=Side(border_style="thin", color="808080"),
            top=Side(border_style="thin", color="808080"),
            bottom=Side(border_style="thin", color="808080")
        )

        if not resultado or "base" not in resultado or "comparables" not in resultado:
            return jsonify({"error": "Entrada inválida: faltan datos de comparación"}), 400

        modelo_dict = {m["codigo_modelo_version"]: m for m in modelos}
        base = modelo_dict.get(resultado["base"])
        comparables = [modelo_dict.get(c["modelo_version"]) for c in resultado.get("comparables", []) if modelo_dict.get(c["modelo_version"])]

        if not base or len(comparables) == 0:
            return jsonify({"error": "Debe existir un modelo base y al menos un comparable válido"}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Comparación"

        firebrick_fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")

        header_cells = [
            (16, 1, "CATEGORIA"),
            (16, 2, "CAMPOS"),
            (16, 3, f"{base['nombre_modelo_comercial']} - {base['nombre_marca']}"),
            (2, 2, f"{base['nombre_modelo_comercial']} - {base['nombre_marca']}"),
            (14, 2, f"Precio: ${base['precio_producto_modelo']}")
        ]

        for i, comp_modelo in enumerate(comparables):
            header_cells.append((16, 4 + i * 2, f"{comp_modelo['nombre_modelo_comercial']} - {comp_modelo['nombre_marca']}"))
            header_cells.append((16, 5 + i * 2, "COMPARATIVO"))
            header_cells.append((2, 4 + i * 2, f"{comp_modelo['nombre_modelo_comercial']} - {comp_modelo['nombre_marca']}"))
            precio = comparables[0].get('precio_producto_modelo')
            if precio is not None:
                col = 4 + i * 2
                cell = ws.cell(row=14, column=col, value=f"Precio: ${precio}")
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for row, col, value in header_cells:
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = firebrick_fill

        def descargar_img(url):
            try:
                if url:
                    res = requests.get(url, timeout=2)
                    if res.ok:
                        return io.BytesIO(res.content)
            except Exception as e:
                print(f"[IMG] Error al descargar: {url} -> {e}")
            return None

        urls = [base.get("path_imagen")] + [
            modelo_dict.get(c["modelo_version"], {}).get("path_imagen")
            for c in resultado.get("comparables", [])
        ]

        with ThreadPoolExecutor() as executor:
            imagenes = list(executor.map(descargar_img, urls))

        if imagenes[0]:
            img = Image(imagenes[0])
            img.width = 250
            img.height = 200
            ws.add_image(img, "B3")

        for i, img_data in enumerate(imagenes[1:]):
            if img_data:
                col = 4 + i * 2
                img = Image(img_data)
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{get_column_letter(col)}3")

        icono_estado = {
            "mejor": {"icono": "👍", "color": "2e7d32"},
            "peor": {"icono": "👎", "color": "d32f2f"},
            "igual": {"icono": "👍👍", "color": "ff9800"},
            "diferente": {"icono": "👍👎", "color": "b300ac"}
        }

        campos_unicos = {}
        for idx, comp in enumerate(resultado.get("comparables", [])):
            detalles_modelo = comp.get("mejor_en", {})
            for categoria, detalles in detalles_modelo.items():
                for det in detalles:
                    clave = (categoria, det["campo"])
                    if clave not in campos_unicos:
                        campos_unicos[clave] = {
                            "categoria": categoria,
                            "campo": det["campo"],
                            "base": det.get("base", ""),
                            "comparables": [("", {"icono": "", "color": "000000"}) for _ in range(len(comparables))]
                        }
                    estado = det.get("estado", "").lower()
                    icono = icono_estado.get(estado, {"icono": "❗", "color": "000000"})
                    campos_unicos[clave]["comparables"][idx] = (det.get("comparable", ""), icono)

        agrupado_por_categoria = defaultdict(list)
        for key in sorted(campos_unicos.keys(), key=lambda k: (k[0], k[1])):
            categoria, campo = key
            agrupado_por_categoria[categoria].append((campo, campos_unicos[key]))

        current_row = 17

        unidades_por_campo = {
            "ALTURA TOTAL": "mm",
            "LONGITUD TOTAL": "mm",
            "ANCHO TOTAL": "mm",
            "PESO SECO": "kg"
        }

        for categoria, campos in agrupado_por_categoria.items():
            inicio_fusion = current_row
            for campo, datos in campos:
                campo_formateado = campo.replace('_', ' ').upper()
                unidad = unidades_por_campo.get(campo_formateado, "")
                valor_base = datos["base"] if datos["base"] not in [None, ""] else "N/A"
                if unidad and valor_base != "N/A":
                    valor_base = f"{valor_base} {unidad}"

                fila = [
                    categoria.upper(),
                    campo_formateado,
                    valor_base
                ]

                for val, icono in datos["comparables"]:

                    val_fmt = val if val not in [None, ""] else "N/A"
                    if unidad and val_fmt != "N/A":
                        val_fmt = f"{val_fmt} {unidad}"
                    fila.append(val_fmt)
                    fila.append(icono["icono"])

                ws.append(fila)
                row_idx = current_row

                for i, (_, icono) in enumerate(datos["comparables"]):
                    col_idx = 5 + i * 2
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(bold=True, color=icono["color"])
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                max_col = 3 + len(comparables) * 2
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = gray_border
                    if col < 4 or col % 2 == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if cell.value is None:
                            cell.value = ""
                current_row += 1

            if len(campos) >= 1:
                ws.merge_cells(start_row=inicio_fusion, start_column=1, end_row=current_row - 1, end_column=1)
                cell = ws.cell(row=inicio_fusion, column=1)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.font = Font(bold=True)

        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.column_dimensions[get_column_letter(2)].width = 25
        ws.column_dimensions[get_column_letter(3)].width = 32
        for i in range(len(comparables)):
            ws.column_dimensions[get_column_letter(4 + i * 2)].width = 32
            ws.column_dimensions[get_column_letter(5 + i * 2)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if output.getbuffer().nbytes < 1000:
            return jsonify({"error": "Archivo generado está vacío o corrupto"}), 500

        return send_file(
            output,
            as_attachment=True,
            download_name="comparacion_modelos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bench_model.route('/get_marcas_por_linea_segmento', methods=['GET'])
@jwt_required()
@cross_origin()
def get_marcas_por_linea_segmento():
    try:
        codigo_linea = request.args.get('codigo_linea', type=int)
        nombre_segmento = request.args.get('nombre_segmento', type=str)
        cil_min = request.args.get('cil_min', type=float)
        cil_max = request.args.get('cil_max', type=float)

        if not codigo_linea or not nombre_segmento:
            return jsonify({"error": "Parámetros 'codigo_linea' y 'nombre_segmento' requeridos"}), 400

        # Subquery para extraer cilindrada como número
        cilindrada_num = func.cast(
            func.nullif(
                func.regexp_replace(Motor.cilindrada_comercial, '[^0-9.]', ''),
                ''
            ),
            Float
        )
        # Subquery: marcas que tengan al menos un modelo válido (por cilindrada o SPARTA)
        subquery = db.session.query(Marca.codigo_marca) \
            .join(ModeloComercial, Marca.codigo_marca == ModeloComercial.codigo_marca) \
            .join(Segmento, (Segmento.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
                            (Segmento.codigo_marca == ModeloComercial.codigo_marca)) \
            .join(ModeloVersion, (ModeloVersion.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
                                  (ModeloVersion.codigo_marca == Marca.codigo_marca)) \
            .join(Motor, ModeloVersion.codigo_motor == Motor.codigo_motor) \
            .filter(Segmento.codigo_linea == codigo_linea) \
            .filter(func.upper(Segmento.nombre_segmento) == func.upper(nombre_segmento.strip())) \
            .filter(Marca.estado_marca == 1)

        if cil_min is not None and cil_max is not None:
            subquery = subquery.filter(
                or_(
                    cilindrada_num.between(cil_min, cil_max),
                    func.upper(ModeloComercial.nombre_modelo) == 'SPARTA 200'
                )
            )

        subquery = subquery.distinct().subquery()

        # Consulta final de marcas válidas
        resultados = db.session.query(Marca.codigo_marca, Marca.nombre_marca) \
            .filter(Marca.codigo_marca.in_(subquery)) \
            .filter(Marca.estado_marca == 1) \
            .distinct() \
            .all()

        marcas = [{
            "codigo_marca": r[0],
            "nombre_marca": r[1]
        } for r in resultados]

        return jsonify(marcas), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bench_model.route('/get_modelos_por_linea_segmento_marca_cilindraje', methods=['GET'])
@jwt_required()
@cross_origin()
def get_modelos_por_linea_segmento_marca_cilindraje():
    try:
        # Obtener parámetros
        codigo_linea = request.args.get('codigo_linea', type=int)
        nombre_segmento = request.args.get('nombre_segmento', type=str)
        codigo_marca = request.args.get('codigo_marca', type=int)
        cil_min = request.args.get('cil_min', type=float)
        cil_max = request.args.get('cil_max', type=float)

        # Validar parámetros obligatorios
        if not codigo_linea or not nombre_segmento or not codigo_marca:
            return jsonify({
                "error": "Parámetros 'codigo_linea', 'nombre_segmento' y 'codigo_marca' requeridos"
            }), 400

        # Subquery para extraer cilindrada como número (segura)
        cilindrada_num = func.cast(
            func.nullif(
                func.regexp_replace(Motor.cilindrada_comercial, '[^0-9.]', ''),
                ''
            ),
            Float
        )

        # Query base con joins y filtros obligatorios
        query = db.session.query(
            ModeloVersion.codigo_modelo_version,
            ModeloVersion.nombre_modelo_version,
            ModeloVersion.anio_modelo_version,
            ModeloVersion.precio_producto_modelo,
            ModeloComercial.nombre_modelo.label('nombre_modelo_comercial'),
            Motor.nombre_motor,
            TipoMotor.nombre_tipo,
            Version.nombre_version,
            Marca.nombre_marca,
            Imagenes.path_imagen,
            Motor.cilindrada_comercial,
        ).join(ModeloComercial,
               (ModeloVersion.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
               (ModeloVersion.codigo_marca == ModeloComercial.codigo_marca)) \
         .join(Segmento,
               (Segmento.codigo_modelo_comercial == ModeloComercial.codigo_modelo_comercial) &
               (Segmento.codigo_marca == ModeloComercial.codigo_marca)) \
         .join(Motor, ModeloVersion.codigo_motor == Motor.codigo_motor) \
         .join(TipoMotor, Motor.codigo_tipo_motor == TipoMotor.codigo_tipo_motor) \
         .join(Version, ModeloVersion.codigo_version == Version.codigo_version) \
         .join(Imagenes, ModeloVersion.codigo_imagen == Imagenes.codigo_imagen) \
         .join(Marca, ModeloVersion.codigo_marca == Marca.codigo_marca) \
         .filter(Segmento.codigo_linea == codigo_linea) \
         .filter(func.upper(Segmento.nombre_segmento) == func.upper(nombre_segmento.strip())) \
         .filter(ModeloVersion.codigo_marca == codigo_marca)

        if cil_min is not None and cil_max is not None:
            filtro_cilindrada = cilindrada_num.between(cil_min, cil_max)

            codigo_sparta = db.session.query(ModeloComercial.codigo_modelo_comercial) \
                .filter(func.upper(ModeloComercial.nombre_modelo) == 'SPARTA 200') \
                .filter(ModeloComercial.codigo_marca == codigo_marca) \
                .scalar()

            if cil_min == 200 and cil_max == 249:
                # Incluir SPARTA 200 explícitamente solo en este rango
                if codigo_sparta:
                    query = query.filter(
                        or_(
                            filtro_cilindrada,
                            ModeloComercial.codigo_modelo_comercial == codigo_sparta
                        )
                    )
                else:
                    query = query.filter(filtro_cilindrada)
            else:
                # Excluir SPARTA 200 de otros rangos, incluso si cae por cilindrada
                if codigo_sparta:
                    query = query.filter(
                        and_(
                            filtro_cilindrada,
                            ModeloComercial.codigo_modelo_comercial != codigo_sparta
                        )
                    )
                else:
                    query = query.filter(filtro_cilindrada)

        resultados = query.all()

        modelos = [{
            "codigo_modelo_version": r[0],
            "nombre_modelo_version": r[1],
            "anio_modelo_version": r[2],
            "precio_producto_modelo": r[3],
            "nombre_modelo_comercial": r[4],
            "nombre_motor": r[5],
            "nombre_tipo": r[6],
            "nombre_version": r[7],
            "nombre_marca": r[8],
            "path_imagen": r[9],
            "cilindrada": r[10]
        } for r in resultados]

        return jsonify(modelos), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#-----------------------------------------------------------------------------------------------------------------------
#-----------------------------------------------------NUEVO-------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------------------

@bench_model.route('/get_modelos_canal', methods=['GET'])
@jwt_required()
@cross_origin()
def get_modelos():
    try:
        nombre_canal = request.args.get('canal', type=str)
        cilindrada = request.args.get('cilindrada', type=str)

        if not nombre_canal or not cilindrada:
            return jsonify({"error": "Los parámetros 'canal' y 'cilindrada' son requeridos."}), 400

        query = text("""
                SELECT MV.CODIGO_MODELO_VERSION,
                       SMC.NOMBRE_MODELO AS NOMBRE_MODELO_COMERCIAL,
                       SM.NOMBRE_MARCA,
                       SS.NOMBRE_SEGMENTO,
                       C.NOMBRE_CANAL,
                       M.CILINDRADA
                FROM ST_MODELO_VERSION mv
                JOIN ST_CLIENTE_CANAL cc
                    ON mv.CODIGO_CLIENTE_CANAL = cc.CODIGO_CLIENTE_CANAL
                    AND mv.CODIGO_MOD_VERS_REPUESTO = cc.CODIGO_MOD_VERS_REPUESTO
                    AND mv.EMPRESA = cc.EMPRESA
                    AND mv.COD_PRODUCTO = cc.COD_PRODUCTO
                JOIN ST_CANAL c
                    ON cc.CODIGO_CANAL = c.CODIGO_CANAL
                JOIN STOCK.ST_MODELO_COMERCIAL SMC
                    ON mv.CODIGO_MODELO_COMERCIAL = SMC.CODIGO_MODELO_COMERCIAL
                JOIN STOCK.ST_MARCA SM
                    ON SMC.CODIGO_MARCA = SM.CODIGO_MARCA
                JOIN STOCK.ST_SEGMENTO SS
                    ON SMC.CODIGO_MODELO_COMERCIAL = SS.CODIGO_MODELO_COMERCIAL
                    AND SMC.CODIGO_MARCA = SS.CODIGO_MARCA
                JOIN STOCK.ST_MOTOR M
                    ON mv.CODIGO_MOTOR = M.CODIGO_MOTOR
                    AND mv.CODIGO_TIPO_MOTOR = M.CODIGO_TIPO_MOTOR
                WHERE c.NOMBRE_CANAL = :canal                  
                  AND REPLACE(UPPER(M.CILINDRADA), ' ', '') LIKE UPPER(CONCAT(:cilindrada, '%'))

            """)

        resultados = db.session.execute(
            query,
            {"canal": nombre_canal.strip(), "cilindrada": f"%{cilindrada.strip()}%"}
        ).fetchall()

        modelos = [{
            "codigo_modelo_version": r[0],
            "nombre_modelo": r[1],
            "nombre_marca": r[2],
            "nombre_segmento": r[3],
            "nombre_canal": r[4],
            "cilindrada": r[5]
        } for r in resultados]

        return jsonify(modelos), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bench_model.route('/get_cilindradas_disponibles', methods=['GET'])
@jwt_required()
@cross_origin()
def get_cilindradas_disponibles():
    try:
        resultados = db.session.query(Motor.cilindrada_comercial).distinct().all()

        # Convertimos a valores numéricos redondeados
        cilindros = []
        for c in resultados:
            try:
                valor = float(str(c[0]).replace('CC', '').strip())
                cilindros.append(valor)
            except:
                continue

        cilindros = sorted(set(cilindros))

        rangos_fijos = [(100, 149), (150, 199), (200, 249), (250, 299), (300, 399), (400, 499), (500, 550)]

        etiquetas_rango = [{
            "label": f"{r_min} CC",
            "min": r_min,
            "max": r_max
        } for r_min, r_max in rangos_fijos]

        return jsonify(etiquetas_rango), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

